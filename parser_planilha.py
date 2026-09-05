"""
parser_planilha.py — Leitura da nova planilha UGR.xlsx
"""
import openpyxl
import re
import unicodedata

def remover_acentos(txt: str) -> str:
    if not txt: return ""
    return ''.join(c for c in unicodedata.normalize('NFD', str(txt)) if unicodedata.category(c) != 'Mn')

def carregar_planilha(path: str) -> dict:
    """Carrega UGR.xlsx e estrutura UGRs, PIs e NDs."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    
    dados = {
        "ugrs": [],
        "nds": [],
        "nd_keywords": []
    }
    
    # Processar aba UNID ADM
    if "UNID ADM" in wb.sheetnames:
        ws = wb["UNID ADM"]
        pi_adm = "VGY01N0105N"
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == "PI" and len(row) > 1 and row[1]:
                pi_adm = str(row[1]).strip()
                break
        
        for row in ws.iter_rows(min_row=2, max_row=45, values_only=True):
            if not row or not row[0]: continue
            sigla = str(row[0]).strip()
            if sigla in ("UNIDADE", "PI") or sigla.startswith("OBS"): continue
            nome = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            ugr = str(row[2]).strip().replace(" ", "") if len(row) > 2 and row[2] else ""
            if ugr:
                dados["ugrs"].append({"sigla": sigla, "nome": nome, "ugr": ugr, "pi": pi_adm})

    # Processar aba UNID ACADÊMICAS
    if " UNID ACADÊMICAS " in wb.sheetnames:
        ws = wb[" UNID ACADÊMICAS "]
        pi_acad = "MGY01N0104N"
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == "PI" and len(row) > 1 and row[1]:
                pi_acad = str(row[1]).strip()
                break
                
        for row in ws.iter_rows(min_row=2, max_row=30, values_only=True):
            if not row or not row[0]: continue
            sigla = str(row[0]).strip()
            if sigla in ("UNIDADE", "PI") or sigla.startswith("OBS"): continue
            nome = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            ugr = str(row[2]).strip().replace(" ", "") if len(row) > 2 and row[2] else ""
            if ugr:
                dados["ugrs"].append({"sigla": sigla, "nome": nome, "ugr": ugr, "pi": pi_acad})

    # Processar aba Elem. de despesa
    if "Elem. de despesa" in wb.sheetnames:
        ws = wb["Elem. de despesa"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row: continue
            
            # Colunas de Principais (índices 0 e 1)
            nd_cod = str(row[0]).strip() if row[0] else ""
            nd_nome = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if nd_cod and nd_cod.isdigit() and len(nd_cod) == 6:
                dados["nds"].append({"nd_cod": nd_cod, "nd_nome": nd_nome})
                
            # Colunas de Outros / Palavras-chave (índices 3 e 4 ou próximos)
            # Na saída anterior as keywords e NDs estavam nas colunas 3 e 4 (index 3 e 4)
            kw = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            kw_nd = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            
            if kw and kw_nd and kw != "None":
                # Limpa códigos de ND que podem ter vindo com formatação errada
                kw_nd_clean = re.sub(r'[^0-9]', '', kw_nd)
                if kw_nd_clean:
                    # Se tiver menos de 6 dígitos, assume que é final (ex: '39' -> '339039')
                    if len(kw_nd_clean) < 6 and kw_nd_clean == "39":
                        kw_nd_clean = "339039"
                    dados["nd_keywords"].append({"keyword": kw, "nd_cod": kw_nd_clean})

    wb.close()
    return dados


def sugerir_celulas(dados_planilha: dict, texto_completo: str, ugr_hint: str = "", nd_hint: str = "", pi_hint: str = "", orig_ugr_hint: str = "") -> dict:
    """Busca a UGR, PI e ND baseado no texto do despacho e nos hints, usando os dicionários carregados do UGR.xlsx"""
    texto_upper = remover_acentos(texto_completo).upper()
    ugr_hint = remover_acentos(ugr_hint).upper() if ugr_hint else ""
    orig_ugr_hint = remover_acentos(orig_ugr_hint).upper() if orig_ugr_hint else ""
    
    def match_ugr(hint, u):
        if not hint: return False
        h = hint
        s = remover_acentos(u["sigla"]).upper()
        n = remover_acentos(u["nome"]).upper()
        
        # Match exato de código UGR ou sigla
        if u["ugr"] == h or h == s:
            return True
        
        # Match quando o hint é o nome completo ou contém o nome completo
        if len(h) > 8 and (h in n or n in h):
            return True
        
        # Palavras do hint que são específicas (não genéricas)
        PALAVRAS_GENERICAS = {
            "DECANATO", "CENTRO", "FACULDADE", "INSTITUTO", "SECRETARIA",
            "DEPARTAMENTO", "UNIVERSIDADE", "DE", "DA", "DO", "E", "EM",
            "GESTAO", "GESTÃO", "PLANEJAMENTO", "ASSESSORIA", "DIRETORIA",
            "COORDENACAO", "COORDENAÇÃO", "DIVISAO", "DIVISÃO",
        }
        # Verifica pedaços como "MATRIZ/DGP" ou palavras
        words = [w for w in re.split(r'[^A-Z0-9]+', h) if len(w) >= 2]
        palavras_nome = set(re.split(r'[^A-Z0-9]+', n))
        
        for w in words:
            # Sigla exata (ex: DGP, DAF)
            if w == s:
                return True
            # Palavra específica do nome que bate exatamente (ex: "PESSOAS" bate em "GESTAO DE PESSOAS")
            if len(w) > 3 and w not in PALAVRAS_GENERICAS and w in palavras_nome:
                return True
        
        return False
    
    ugr_encontrada = None
    motivo_ugr = "Nenhuma UGR encontrada no corpo do despacho."
    
    texto_limpo = re.sub(r'[^A-Z0-9]+', ' ', texto_upper)
    palavras_texto = set(texto_limpo.split())
    
    # 0. Tenta o ugr_hint PRIMEIRO (pois ele vem da fonte de crédito, que é preenchida pela IA se ativada)
    if ugr_hint:
        for u in dados_planilha["ugrs"]:
            if match_ugr(ugr_hint, u):
                ugr_encontrada = u
                motivo_ugr = f"UGR sugerida através da inteligência semântica / metadados ('{ugr_hint}')."
                break

    # 1. Tenta varrer todo o texto (Desabilitado a pedido do usuário: somente a IA deve decidir)
    # 2. Se não achou no texto principal, tenta usar o hint do centro de custo (Desabilitado)

                
    if not ugr_encontrada and orig_ugr_hint:
        for u in dados_planilha["ugrs"]:
            if match_ugr(orig_ugr_hint, u):
                ugr_encontrada = u
                motivo_ugr = f"Encontrado através da dica original '{orig_ugr_hint}'."
                break
                
    # 3. Se ainda não achou, tenta palavras-chave importantes do nome no hint
    if not ugr_encontrada and ugr_hint:
        for u in dados_planilha["ugrs"]:
            n = remover_acentos(u["nome"]).upper()
            if n:
                n_words = [w for w in re.split(r'[^A-Z0-9]+', n) if len(w) > 4 and w not in ("DECANATO", "CENTRO", "FACULDADE", "INSTITUTO", "SECRETARIA", "DEPARTAMENTO", "UNIVERSIDADE")]
                if n_words:
                    h_words = set(re.split(r'[^A-Z0-9]+', ugr_hint))
                    for nw in n_words:
                        if nw in h_words:
                            ugr_encontrada = u
                            motivo_ugr = f"A palavra-chave '{nw}' do nome da UGR foi encontrada no centro de custo."
                            break
                    if ugr_encontrada:
                        break
    nd_encontrada = nd_hint
    motivo_nd = f"ND herdada da semântica ({nd_hint})." if nd_hint else "Nenhuma ND identificada."
    
    if not nd_encontrada or nd_encontrada == "339000":
        for kw_item in dados_planilha["nd_keywords"]:
            if remover_acentos(kw_item["keyword"]).upper() in texto_upper:
                nd_encontrada = kw_item["nd_cod"]
                motivo_nd = f"A palavra-chave '{kw_item['keyword']}' foi encontrada no corpo do texto."
                break
                
    # 3. Montar resultado
    ugr_cod  = ugr_encontrada["ugr"]  if ugr_encontrada else ""
    ugr_nome = ugr_encontrada["nome"] if ugr_encontrada else "(preencha a UGR)"
    ugr_pi   = ugr_encontrada["pi"]   if ugr_encontrada else ""
    
    origem = {
        "esfera": "1",
        "ptres": "230639",
        "fonte": "1050A000AP",
        "nd": "339000",
        "nd_nome": "APLICACOES DIRETAS",
        "ugr": ugr_cod,
        "ugr_nome": ugr_nome,
        "pi": ugr_pi,
        "pi_nome": "",
        "saldo": 0,
        "valor": ""
    }
        
    nome_nd_destino = "(preencha o ND específico)"
    if nd_encontrada and nd_encontrada != "339000":
        nome_nd_destino = f"Natureza {nd_encontrada}"
        # Busca nome oficial da ND se existir
        for nd in dados_planilha["nds"]:
            if nd["nd_cod"] == nd_encontrada:
                nome_nd_destino = nd["nd_nome"]
                break
                
    destino = {
        **origem,
        "nd": nd_encontrada if nd_encontrada and nd_encontrada != "339000" else "",
        "nd_nome": nome_nd_destino
    }
        
    return {
        "origem": origem,
        "destino": destino,
        "alternativas": [],
        "total_acao": 1,
        "avisos": [],
        "encontrou": ugr_encontrada is not None,
        "prova_noves": {
            "ugr": motivo_ugr,
            "nd": motivo_nd
        }
    }
